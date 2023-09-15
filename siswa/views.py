from django.shortcuts import redirect, render
from django.views.decorators.csrf import csrf_exempt
from django.contrib.auth.hashers import make_password, check_password
from django.contrib.auth import authenticate, login
from django.contrib.auth.models import User
from django.contrib import auth
import json
from django.http.response import HttpResponse
import qrcode
import base64
import os
from io import BytesIO
from django.urls import reverse
from app_siswa.settings import MEDIA_ROOT
from siswa.forms import GuruForm, PelanggaranForm, SiswaPelanggaranForm, TindakanForm
from siswa.models import Pelanggaran, Siswa, Guru, SiswaPelanggaran, Tindakan
from django.db import connection
from datetime import datetime
from django.db.models import Q,Sum

#pdf convert
from django.template.loader import get_template
from xhtml2pdf import pisa

#import excel
import openpyxl
import xlsxwriter
import io

# Create your views here.
def index(request):
    return render(request, "login.html")


def login(request):
    return render(request, "login.html")

def loginpelanggaran(request):
    return render(request, "loginpelanggaran.html",{'nomor':request.session.get("nomorinduk")})

@csrf_exempt
def auth(request):
    username = request.POST["username"]
    password = request.POST["password"]
    user = Guru.objects.filter(username=username).filter(password=password).all()
    if user != None:
        for val in user.iterator():
            request.session["nama"] = val.nama_lengkap
            request.session["nuptk"] = val.nuptk
            request.session["logged_in"] = True
        arr = json.dumps({"success": 1, "message": "Success Login"})
    else:
        arr = json.dumps({"success": 0, "message": "Username & Password Tidak Cocok"})
    return HttpResponse(arr)


def dashboard(request):
    if request.session.get("logged_in") != True:
        return redirect("login")
    siswa = Siswa.objects.count()
    kategori = Pelanggaran.objects.count()
    pelanggaran = SiswaPelanggaran.objects.count()
    with connection.cursor() as con:
        con.execute("SELECT COUNT(s.siswa_id) as total FROM siswa_siswapelanggaran s INNER JOIN siswa_pelanggaran p ON s.pelanggaran_id=p.id_pelanggaran where p.point_pelanggaran >=20 AND p.point_pelanggaran <30")
        rows = con.fetchone()
        bk = rows[0]
    with connection.cursor() as con:
        con.execute("SELECT COUNT(s.siswa_id) as total FROM siswa_siswapelanggaran s INNER JOIN siswa_pelanggaran p ON s.pelanggaran_id=p.id_pelanggaran where p.point_pelanggaran >=90 AND p.point_pelanggaran <100")
        rows = con.fetchone()
        ortu = rows[0]
    with connection.cursor() as con:
        con.execute("SELECT COUNT(s.siswa_id) as total FROM siswa_siswapelanggaran s INNER JOIN siswa_pelanggaran p ON s.pelanggaran_id=p.id_pelanggaran where p.point_pelanggaran=100")
        rows = con.fetchone()
        keluar = rows[0]
    output ={
        "url": "dashboard",
        'siswa':siswa,
        'kategori':kategori,
        'pelanggaran':pelanggaran,
        'bk':str(bk),
        'ortu':str(ortu),
        'keluar':str(keluar),
        }
    return render(request, "dashboard.html", output)


def logout(request):
    request.session.clear()
    return redirect("login")


def siswa(request):
    if request.session.get("logged_in") != True:
        return redirect("login")
    return render(request, "siswa.html", {"url": "siswa"})

@csrf_exempt
def svsiswa(request):
    if request.method == "POST":
        urlmap= 'http://192.168.72.222:8000/addpelanggaran/'
        pk = request.POST.get("id")
        nama = request.POST.get("nama")
        tempat = request.POST.get("tempat")
        tgl = request.POST.get("tgl_lahir")
        gender = request.POST.get("gender")
        lembaga = request.POST.get("lembaga")
        nomor = request.POST.get("nomor")
        jurusan = request.POST.get("jurusan")
        kelas = request.POST.get("kelas")
        try:
            if pk != None:
                cek = Siswa.objects.filter(id_siswa=pk).all()
                for v in cek.iterator():
                    induk = v.nomor_induk
                    if induk != nomor:
                        os.unlink(MEDIA_ROOT + "/siswa/" + str(v.qr_code))
                        qr_code_url = urlmap+nomor
                        qr_code = qrcode.QRCode(
                            version=1,
                            error_correction=qrcode.constants.ERROR_CORRECT_L,
                            box_size=10,
                            border=4,
                        )
                        qr_code.add_data(qr_code_url)
                        qr_code.make(fit=True)
                        qr_code_image = qr_code.make_image(
                            fill_color="black", back_color="white"
                        )

                        qr_code_buffer = BytesIO()
                        qr_code_image.save(qr_code_buffer, format="PNG")

                        qr_code_directory = os.path.join("media", "siswa")
                        os.makedirs(qr_code_directory, exist_ok=True)

                        qr_code_filepath = os.path.join(
                            qr_code_directory, f"qr_code_{nomor}.png"
                        )

                        with open(qr_code_filepath, "wb") as f:
                            f.write(qr_code_buffer.getvalue())
                    editdata = Siswa.objects.get(id_siswa=pk)
                    editdata.nama_lengkap = nama
                    editdata.tempat_lahir = tempat
                    editdata.tanggal_lahir = tgl
                    editdata.jenis_kelamin = gender
                    editdata.lembaga = lembaga
                    editdata.nomor_induk = nomor
                    editdata.jurusan = jurusan
                    editdata.kelas = kelas
                    editdata.qr_code = f"qr_code_{nomor}.png"
                    editdata.save()
                else:
                    editdata = Siswa.objects.get(id_siswa=pk)
                    editdata.nama_lengkap = nama
                    editdata.tempat_lahir = tempat
                    editdata.tanggal_lahir = tgl
                    editdata.jenis_kelamin = gender
                    editdata.lembaga = lembaga
                    editdata.nomor_induk = nomor
                    editdata.jurusan = jurusan
                    editdata.kelas = kelas
                    editdata.save()

            else:
                qr_code_url = urlmap+nomor
                qr_code = qrcode.QRCode(
                    version=1,
                    error_correction=qrcode.constants.ERROR_CORRECT_L,
                    box_size=10,
                    border=4,
                )
                qr_code.add_data(qr_code_url)
                qr_code.make(fit=True)
                qr_code_image = qr_code.make_image(
                    fill_color="black", back_color="white"
                )

                qr_code_buffer = BytesIO()
                qr_code_image.save(qr_code_buffer, format="PNG")

                qr_code_directory = os.path.join("media", "siswa")
                os.makedirs(qr_code_directory, exist_ok=True)

                qr_code_filepath = os.path.join(
                    qr_code_directory, f"qr_code_{nomor}.png"
                )

                with open(qr_code_filepath, "wb") as f:
                    f.write(qr_code_buffer.getvalue())

                Siswa.objects.create(
                    nama_lengkap=nama,
                    tempat_lahir=tempat,
                    tanggal_lahir=tgl,
                    jenis_kelamin=gender,
                    lembaga=lembaga,
                    nomor_induk=nomor,
                    jurusan=jurusan,
                    kelas=kelas,
                    qr_code=f"qr_code_{nomor}.png",
                )
            arr = json.dumps({"success": 1, "message": "Tersimpan"})
        except IndexError:
            arr = json.dumps({"success": 0, "message": "Failed"})
    else:
        arr = json.dumps({"success": 0, "message": "Failed"})
    return HttpResponse(arr)


@csrf_exempt
def vsiswa(request):
    data = []
    out = []
    no = 1
    rows = Siswa.objects.order_by("nama_lengkap").all()
    for val in rows.iterator():
        nom = "<center>" + str(no) + "</center>"
        nomor = val.nomor_induk
        alamat = val.alamat.capitalize()
        nama = val.nama_lengkap.capitalize()
        tempat = val.tempat_lahir.capitalize()
        tgl = val.tanggal_lahir.strftime("%d/%m/%Y")
        if val.jenis_kelamin == "L":
            gender = "Laki-laki"
        else:
            gender = "Perempuan"

        lembaga = val.lembaga
        jurusan = val.jurusan
        kelas = "Kelas " + val.kelas
        rombel = val.rombel
        barcode = "<a href='media/siswa/" + str(val.qr_code) + "' target='_blank'><img src='media/siswa/" + str(val.qr_code) + "' height='100'/></a>"

        button = (
            '<a href="javascript:detail('
            + str(val.id_siswa)
            + ')" title = "Edit details" class = "btn btn-sm btn-warning btn-icon btn-icon-md" >'
        )
        button += "<i class='bx bx-edit-alt' ></i>"
        button += "</a>&nbsp;"
        button += (
            '<a href="javascript:hapus('
            + str(val.id_siswa)
            + ')" title = "Delete" class = "btn btn-sm btn-danger btn-icon btn-icon-md" >'
        )
        button += "<i class='bx bx-message-alt-x'></i>"
        button += "</a>"

        out = [
            nom,
            nomor,
            nama,
            tempat,
            tgl,
            gender,
            alamat,
            jurusan,
            kelas,
            barcode,
            button,
        ]
        data.append(out)
        no += 1

    output = json.dumps({"data": data})
    return HttpResponse(output, content_type="application/json")


@csrf_exempt
def dsiswa(request):
    id = request.POST.get("id")
    dos = Siswa.objects.filter(id_siswa=id).all()
    data = []
    value = {}
    for val in dos.iterator():
        value = {
            "nama": val.nama_lengkap,
            "id": val.id_siswa,
            "gender": val.jenis_kelamin,
            "tempat": val.tempat_lahir,
            "tanggal": val.tanggal_lahir.strftime("%d/%m/%Y"),
            "gender": val.jenis_kelamin,
            "lembaga": val.lembaga,
            "jurusan": val.jurusan,
            "kelas": val.kelas,
            "nomor": val.nomor_induk,
            "alamat": val.alamat,
            "barcode": str(val.qr_code),
        }
        data.append(value)
    output = json.dumps({"data": data})
    return HttpResponse(output, content_type="application/json")


@csrf_exempt
def hsiswa(request):
    if request.method == "POST":
        pk = request.POST.get("id")
        cek = Siswa.objects.filter(id_siswa=pk).all()
        for v in cek.iterator():
            os.unlink(MEDIA_ROOT + "/siswa/" + str(v.qr_code))
        data = Siswa.objects.get(id_siswa=pk)
        sql = data.delete()
        if sql:
            arr = json.dumps({"success": 1, "message": "Terhapus"})
        else:
            arr = json.dumps({"success": 0, "message": "Failed"})
    else:
        arr = json.dumps({"success": 0, "message": "Failed"})
    return HttpResponse(arr)


def petugas(request):
    if request.session.get("logged_in") != True:
        return redirect("login")
    return render(request, "petugas.html", {"url": "petugas"})


@csrf_exempt
def svpetugas(request):
    if request.method == "POST":
        pk = request.POST.get("id_guru")
        if pk != "":
            editdata = Guru.objects.get(id_guru=pk)
            save = GuruForm(request.POST or None, instance=editdata)
        else:
            save = GuruForm(request.POST or None)

        if save.is_valid:
            sql = save.save()
            if sql:
                arr = json.dumps({"success": 1, "message": "Tersimpan"})
            else:
                arr = json.dumps({"success": 0, "message": "Failed"})
        else:
            arr = json.dumps({"success": 0, "message": "Failed"})
    else:
        arr = json.dumps({"success": 0, "message": "Failed"})
    return HttpResponse(arr)


@csrf_exempt
def vpetugas(request):
    data = []
    out = []
    no = 1
    rows = Guru.objects.filter(~Q(username='admin')).order_by("nama_lengkap").all()
    for val in rows.iterator():
        nom = "<center>" + str(no) + "</center>"
        nuptk = val.nuptk
        nama = val.nama_lengkap.upper()
        tgl = val.tanggal_lahir.strftime("%d/%m/%Y")
        kontak = val.nomer_hp
        username = val.username
        passwrod = "************"
        button = (
            '<a href="javascript:detail('
            + str(val.id_guru)
            + ')" title = "Edit details" class = "btn btn-sm btn-warning btn-icon btn-icon-md" >'
        )
        button += "<i class='bx bx-edit-alt' ></i>"
        button += "</a>&nbsp;"
        button += (
            '<a href="javascript:hapus('
            + str(val.id_guru)
            + ')" title = "Delete" class = "btn btn-sm btn-danger btn-icon btn-icon-md" >'
        )
        button += "<i class='bx bx-message-alt-x'></i>"
        button += "</a>"

        out = [
            nom,
            nuptk,
            nama,
            tgl,
            kontak,
            username,
            passwrod,
            button,
        ]
        data.append(out)
        no += 1

    output = json.dumps({"data": data})
    return HttpResponse(output, content_type="application/json")


@csrf_exempt
def dpetugas(request):
    id = request.POST.get("id")
    dos = Guru.objects.filter(id_guru=id).all()
    data = []
    value = {}
    for val in dos.iterator():
        value = {
            "nama": val.nama_lengkap,
            "id": val.id_guru,
            "tanggal": val.tanggal_lahir.strftime("%d/%m/%Y"),
            "nuptk": val.nuptk,
            "kontak": val.nomer_hp,
            "username": val.username,
            "password": val.password,
        }
        data.append(value)
    output = json.dumps({"data": data})
    return HttpResponse(output, content_type="application/json")


@csrf_exempt
def hpetugas(request):
    if request.method == "POST":
        pk = request.POST.get("id")
        data = Guru.objects.get(id_guru=pk)
        sql = data.delete()
        if sql:
            arr = json.dumps({"success": 1, "message": "Terhapus"})
        else:
            arr = json.dumps({"success": 0, "message": "Failed"})
    else:
        arr = json.dumps({"success": 0, "message": "Failed"})
    return HttpResponse(arr)


def kategori(request):
    if request.session.get("logged_in") != True:
        return redirect("login")
    return render(request, "kategori.html", {"url": "kategori"})


@csrf_exempt
def svkategori(request):
    if request.method == "POST":
        pk = request.POST.get("id")
        if pk != "":
            editdata = Pelanggaran.objects.get(id_pelanggaran=pk)
            save = PelanggaranForm(request.POST or None, instance=editdata)
        else:
            save = PelanggaranForm(request.POST or None)

        if save.is_valid:
            sql = save.save()
            if sql:
                arr = json.dumps({"success": 1, "message": "Tersimpan"})
            else:
                arr = json.dumps({"success": 0, "message": "Failed"})
        else:
            arr = json.dumps({"success": 0, "message": "Failed"})
    else:
        arr = json.dumps({"success": 0, "message": "Failed"})
    return HttpResponse(arr)


@csrf_exempt
def vkategori(request):
    data = []
    out = []
    no = 1
    rows = Pelanggaran.objects.order_by("point_pelanggaran").all()
    for val in rows.iterator():
        nom = "<center>" + str(no) + "</center>"
        nama = val.nama_pelanggaran.upper()
        kategori = val.kategori.upper()
        point = str(val.point_pelanggaran) + " Point"
        button = (
            '<a href="javascript:detail('
            + str(val.id_pelanggaran)
            + ')" title = "Edit details" class = "btn btn-sm btn-warning btn-icon btn-icon-md" >'
        )
        button += "<i class='bx bx-edit-alt' ></i>"
        button += "</a>&nbsp;"
        button += (
            '<a href="javascript:hapus('
            + str(val.id_pelanggaran)
            + ')" title = "Delete" class = "btn btn-sm btn-danger btn-icon btn-icon-md" >'
        )
        button += "<i class='bx bx-message-alt-x'></i>"
        button += "</a>"

        out = [
            nom,
            nama,
            kategori,
            point,
            button,
        ]
        data.append(out)
        no += 1

    output = json.dumps({"data": data})
    return HttpResponse(output, content_type="application/json")


@csrf_exempt
def dkategori(request):
    id = request.POST.get("id")
    dos = Pelanggaran.objects.filter(id_pelanggaran=id).all()
    data = []
    value = {}
    for val in dos.iterator():
        value = {
            "nama": val.nama_pelanggaran,
            "id": str(val.id_pelanggaran),
            "point": str(val.point_pelanggaran),
            "kategori": val.kategori,
        }
        data.append(value)
    output = json.dumps({"data": data})
    return HttpResponse(output, content_type="application/json")


@csrf_exempt
def hkategori(request):
    if request.method == "POST":
        pk = request.POST.get("id")
        data = Pelanggaran.objects.get(id_pelanggaran=pk)
        sql = data.delete()
        if sql:
            arr = json.dumps({"success": 1, "message": "Terhapus"})
        else:
            arr = json.dumps({"success": 0, "message": "Failed"})
    else:
        arr = json.dumps({"success": 0, "message": "Failed"})
    return HttpResponse(arr)


def sanksi(request):
    if request.session.get("logged_in") != True:
        return redirect("login")
    return render(request, "sanksi.html", {"url": "sanksi"})


@csrf_exempt
def svsanksi(request):
    if request.method == "POST":
        pk = request.POST.get("id")
        if pk != "":
            editdata = Tindakan.objects.get(id_tindakan=pk)
            save = TindakanForm(request.POST or None, instance=editdata)
        else:
            save = TindakanForm(request.POST or None)

        if save.is_valid:
            sql = save.save()
            if sql:
                arr = json.dumps({"success": 1, "message": "Tersimpan"})
            else:
                arr = json.dumps({"success": 0, "message": "Failed"})
        else:
            arr = json.dumps({"success": 0, "message": "Failed"})
    else:
        arr = json.dumps({"success": 0, "message": "Failed"})
    return HttpResponse(arr)


@csrf_exempt
def vsanksi(request):
    data = []
    out = []
    no = 1
    rows = Tindakan.objects.order_by("range_point").all()
    for val in rows.iterator():
        nom = "<center>" + str(no) + "</center>"
        nama = val.tindakan.upper()
        point = str(val.range_point) + " Point"
        button = (
            '<a href="javascript:detail('
            + str(val.id_tindakan)
            + ')" title = "Edit details" class = "btn btn-sm btn-warning btn-icon btn-icon-md" >'
        )
        button += "<i class='bx bx-edit-alt' ></i>"
        button += "</a>&nbsp;"
        button += (
            '<a href="javascript:hapus('
            + str(val.id_tindakan)
            + ')" title = "Delete" class = "btn btn-sm btn-danger btn-icon btn-icon-md" >'
        )
        button += "<i class='bx bx-message-alt-x'></i>"
        button += "</a>"

        out = [
            nom,
            nama,
            point,
            button,
        ]
        data.append(out)
        no += 1

    output = json.dumps({"data": data})
    return HttpResponse(output, content_type="application/json")


@csrf_exempt
def dsanksi(request):
    id = request.POST.get("id")
    dos = Tindakan.objects.filter(id_tindakan=id).all()
    data = []
    value = {}
    for val in dos.iterator():
        value = {
            "tindakan": val.tindakan,
            "id": val.id_tindakan,
            "point": str(val.range_point),
        }
        data.append(value)
    output = json.dumps({"data": data})
    return HttpResponse(output, content_type="application/json")


@csrf_exempt
def hsanksi(request):
    if request.method == "POST":
        pk = request.POST.get("id")
        data = Tindakan.objects.get(id_tindakan=pk)
        sql = data.delete()
        if sql:
            arr = json.dumps({"success": 1, "message": "Terhapus"})
        else:
            arr = json.dumps({"success": 0, "message": "Failed"})
    else:
        arr = json.dumps({"success": 0, "message": "Failed"})
    return HttpResponse(arr)


def pelanggaran(request):
    if request.session.get("logged_in") != True:
        return redirect("login")
    return render(request, "pelanggaran.html", {"url": "pelanggaran"})

def addpelanggaran(request,nomorinduk):
    if request.session.get("logged_in") != True:
        request.session["nomorinduk"] = nomorinduk
        return redirect("loginpelanggaran")
    sql = Siswa.objects.filter(nomor_induk=nomorinduk).all()
    for v in sql.iterator():
        ttl = SiswaPelanggaran.objects.filter(siswa_id=v.id_siswa).all()
        total = 0
        for vt in ttl.iterator():
            pel = Pelanggaran.objects.filter(id_pelanggaran=vt.pelanggaran_id).all()
            for t in pel.iterator():
                total += t.point_pelanggaran
             
        output= {
            'id':str(v.id_siswa),
            'nomor':v.nomor_induk,
            'nama':v.nama_lengkap,
            'tgl':v.tanggal_lahir.strftime('%d/%m/%Y'),
            'kelas':'Kelas '+v.kelas,
            'jurusan':v.jurusan,
            'point': str(total)+' Point',
            }
     
    return render(request, "addpelanggaran.html",output)

@csrf_exempt
def listpelanggaran(request):
    sql = Pelanggaran.objects.all()
    data = ""
    for val in sql.iterator():
        data += (
            '<option value="'
            + str(val.id_pelanggaran)
            + '">'
            + val.nama_pelanggaran.upper()
            + " ("
            + str(val.point_pelanggaran)
            + " Point)"
            + "</option>"
        )
    return HttpResponse(data)


@csrf_exempt
def listsiswa(request):
    sql = Siswa.objects.order_by("nama_lengkap").all()
    data = ""
    for val in sql.iterator():
        data += (
            '<option value="'
            + str(val.id_siswa)
            + '">'
            + val.nama_lengkap.upper()
            + "</option>"
        )
    return HttpResponse(data)


@csrf_exempt
def svpelanggaran(request):
    if request.method == "POST":
        pk = request.POST.get("id")
        if pk != "":
            editdata = SiswaPelanggaran.objects.get(id_pelanggaran=pk)
            save = SiswaPelanggaranForm(request.POST or None, instance=editdata)
        else:
            save = SiswaPelanggaranForm(request.POST or None)

        if save.is_valid:
            sql = save.save()
            if sql:
                arr = json.dumps({"success": 1, "message": "Tersimpan"})
            else:
                arr = json.dumps({"success": 0, "message": "Failed"})
        else:
            arr = json.dumps({"success": 0, "message": "Failed"})
    else:
        arr = json.dumps({"success": 0, "message": "Failed"})
    return HttpResponse(arr)


@csrf_exempt
def vpelanggaran(request):
    data = []
    out = []
    no = 1
    rows = SiswaPelanggaran.objects.order_by("-tanggal_pelanggaran").all()
    for val in rows.iterator():
        nom = "<center>" + str(no) + "</center>"
        sis = Siswa.objects.filter(id_siswa=val.siswa_id).all()
        for v in sis.iterator():
            siswa = v.nama_lengkap.upper()
            kelas = v.kelas
            induk = v.nomor_induk
        pel = Pelanggaran.objects.filter(id_pelanggaran=val.pelanggaran_id).all()
        for v in pel.iterator():
            pelanggaran = v.nama_pelanggaran.upper()
            point = str(v.point_pelanggaran) + " Point"
            status = v.kategori.upper()
        ttl = SiswaPelanggaran.objects.filter(siswa_id=val.siswa_id).all()
        total = 0
        for v in ttl.iterator():
            pel = Pelanggaran.objects.filter(id_pelanggaran=v.pelanggaran_id).all()
            for t in pel.iterator():
                total += t.point_pelanggaran
        with connection.cursor() as con:
            con.execute("SELECT tindakan FROM siswa_tindakan WHERE range_point <=%s ORDER BY range_point DESC LIMIT 1",[total])
            rows = con.fetchall()
            if rows != None:
                for v in rows:
                    tindakan = v[0].upper()
            else:
                tindakan = "-"
                   
        tanggal = val.tanggal_pelanggaran.strftime("%d/%m/%Y, %H:%M")
        button = (
            '<a href="javascript:detail('
            + str(val.id_pelanggaran)
            + ')" title = "Edit details" class = "btn btn-sm btn-warning btn-icon btn-icon-md" >'
        )
        button += "<i class='bx bx-edit-alt' ></i>"
        button += "</a>&nbsp;"
        button += (
            '<a href="javascript:hapus('
            + str(val.id_pelanggaran)
            + ')" title = "Delete" class = "btn btn-sm btn-danger btn-icon btn-icon-md" >'
        )
        button += "<i class='bx bx-message-alt-x'></i>"
        button += "</a>&nbsp;"
        button += (
            '<a href="javascript:pemanggilan('
            + str(val.siswa_id)
            + ')" title = "Form Pemanggilan" class = "btn btn-sm btn-success btn-icon btn-icon-md" >'
        )
        button += "<i class='bx bx-printer'></i></a>&nbsp;"
        button += (
            '<a href="javascript:pengeluaran('
            + str(val.siswa_id)
            + ')" title = "Form Pengeluaran" class = "btn btn-sm btn-info btn-icon btn-icon-md" >'
        )
        button += "<i class='bx bx-printer'></i>"
        button += "</a>"

        out = [
            nom,
            induk,
            siswa,
            kelas,
            pelanggaran,
            point,
            status,
            str(total) + " Point",
            tindakan,
            tanggal,
            button,
        ]
        data.append(out)
        no += 1

    output = json.dumps({"data": data})
    return HttpResponse(output, content_type="application/json")


@csrf_exempt
def dtlvpelanggaran(request):
    id = request.POST.get("id")
    data = []
    out = []
    no = 1
    rows = SiswaPelanggaran.objects.filter(siswa_id=id).order_by("-tanggal_pelanggaran").all()
    for val in rows.iterator():
        nom = "<center>" + str(no) + "</center>"
        sis = Siswa.objects.filter(id_siswa=val.siswa_id).all()
        for v in sis.iterator():
            siswa = v.nama_lengkap.upper()
            kelas = v.kelas
            induk = v.nomor_induk
        pel = Pelanggaran.objects.filter(id_pelanggaran=val.pelanggaran_id).all()
        for v in pel.iterator():
            pelanggaran = v.nama_pelanggaran.upper()
            point = str(v.point_pelanggaran) + " Point"
            status = v.kategori.upper()
        ttl = SiswaPelanggaran.objects.filter(siswa_id=val.siswa_id).all()
        total = 0
        for v in ttl.iterator():
            pel = Pelanggaran.objects.filter(id_pelanggaran=v.pelanggaran_id).all()
            for t in pel.iterator():
                total += t.point_pelanggaran
        with connection.cursor() as con:
            con.execute("SELECT tindakan FROM siswa_tindakan WHERE range_point <=%s ORDER BY range_point DESC LIMIT 1",[total])
            rows = con.fetchall()
            if rows != None:
                for v in rows:
                    tindakan = v[0].upper()
            else:
                tindakan = "-"
                   
        tanggal = val.tanggal_pelanggaran.strftime("%d/%m/%Y, %H:%M")
        button = (
            '<a href="javascript:detail('
            + str(val.id_pelanggaran)
            + ')" title = "Edit details" class = "btn btn-sm btn-warning btn-icon btn-icon-md" >'
        )
        button += "<i class='bx bx-edit-alt' ></i>"
        button += "</a>&nbsp;"
        button += (
            '<a href="javascript:hapus('
            + str(val.id_pelanggaran)
            + ')" title = "Delete" class = "btn btn-sm btn-danger btn-icon btn-icon-md" >'
        )
        button += "<i class='bx bx-message-alt-x'></i>"
        button += "</a>&nbsp;"
        button += (
            '<a href="javascript:pemanggilan('
            + str(val.siswa_id)
            + ')" title = "Form Pemanggilan" class = "btn btn-sm btn-success btn-icon btn-icon-md" >'
        )
        button += "<i class='bx bx-printer'></i></a>&nbsp;"
        button += (
            '<a href="javascript:pengeluaran('
            + str(val.siswa_id)
            + ')" title = "Form Pengeluaran" class = "btn btn-sm btn-info btn-icon btn-icon-md" >'
        )
        button += "<i class='bx bx-printer'></i>"
        button += "</a>"

        out = [
            nom,
            induk,
            siswa,
            kelas,
            pelanggaran,
            point,
            status,
            str(total) + " Point",
            tindakan,
            tanggal,
            button,
        ]
        data.append(out)
        no += 1

    output = json.dumps({"data": data})
    return HttpResponse(output, content_type="application/json")

@csrf_exempt
def dpelanggaran(request):
    id = request.POST.get("id")
    dos = SiswaPelanggaran.objects.filter(id_pelanggaran=id).all()
    data = []
    value = {}
    for val in dos.iterator():
        value = {
            "pelanggaran": str(val.pelanggaran_id),
            "id": val.id_pelanggaran,
            "siswa": str(val.siswa_id),
            "tgl": str(val.tanggal_pelanggaran),
            "jam": val.tanggal_pelanggaran.strftime("%H:%M:%S"),
        }
        data.append(value)
    output = json.dumps({"data": data})
    return HttpResponse(output, content_type="application/json")


@csrf_exempt
def hpelanggaran(request):
    if request.method == "POST":
        pk = request.POST.get("id")
        data = SiswaPelanggaran.objects.get(id_pelanggaran=pk)
        sql = data.delete()
        if sql:
            arr = json.dumps({"success": 1, "message": "Terhapus"})
        else:
            arr = json.dumps({"success": 0, "message": "Failed"})
    else:
        arr = json.dumps({"success": 0, "message": "Failed"})
    return HttpResponse(arr)


def rekapan(request):
    if request.session.get("logged_in") != True:
        return redirect("login")
    return render(request, "rekapan.html", {"url": "rekapan"})


@csrf_exempt
def vrekapan(request):
    data = []
    out = []
    no = 1
    rows = SiswaPelanggaran.objects.order_by("-tanggal_pelanggaran").all()
    for val in rows.iterator():
        nom = "<center>" + str(no) + "</center>"
        sis = Siswa.objects.filter(id_siswa=val.siswa_id).all()
        for v in sis.iterator():
            siswa = v.nama_lengkap.upper()
            kelas = v.kelas
            induk = v.nomor_induk
        pel = Pelanggaran.objects.filter(id_pelanggaran=val.pelanggaran_id).all()
        for v in pel.iterator():
            pelanggaran = v.nama_pelanggaran.upper()
            point = str(v.point_pelanggaran) + " Point"
            status = v.kategori.upper()
        ttl = SiswaPelanggaran.objects.filter(siswa_id=val.siswa_id).all()
        total = 0
        for v in ttl.iterator():
            pel = Pelanggaran.objects.filter(id_pelanggaran=v.pelanggaran_id).all()
            for t in pel.iterator():
                total += t.point_pelanggaran
        with connection.cursor() as con:
            con.execute("SELECT tindakan FROM siswa_tindakan WHERE range_point <=%s ORDER BY range_point DESC LIMIT 1",[total])
            rows = con.fetchall()
            if rows != None:
                for v in rows:
                    tindakan = v[0].upper()
            else:
                tindakan = "-"
                   
        tanggal = val.tanggal_pelanggaran.strftime("%d/%m/%Y, %H:%M")
        button = (
            '<a href="javascript:pemanggilan('
            + str(val.siswa_id)
            + ')" title = "Form Pemanggilan" class = "btn btn-sm btn-success btn-icon btn-icon-md" >'
        )
        button += "<i class='bx bx-printer'></i></a>&nbsp;"
        button += (
            '<a href="javascript:pengeluaran('
            + str(val.siswa_id)
            + ')" title = "Form Pengeluaran" class = "btn btn-sm btn-info btn-icon btn-icon-md" >'
        )
        button += "<i class='bx bx-printer'></i>"
        button += "</a>"

        out = [
            nom,
            induk,
            siswa,
            kelas,
            pelanggaran,
            point,
            status,
            str(total) + " Point",
            tindakan,
            tanggal,
            button,
        ]
        data.append(out)
        no += 1

    output = json.dumps({"data": data})
    return HttpResponse(output, content_type="application/json")



def pemanggilan(request):
    if request.session.get("logged_in") != True:
        return redirect("login")
    template_name = "formpemanggilan.html"
    siswa = request.GET['siswa']
    tgl = request.GET['tgl']
    date_format = '%Y-%m-%d %H:%M:%S'
    date_obj = datetime.strptime(tgl, date_format)
    records = Siswa.objects.filter(id_siswa=siswa).all()
    day = date_obj.strftime("%A")
    tanggal = date_obj.strftime("%d")
    bln = date_obj.strftime("%m")
    tahun = date_obj.strftime("%Y")
    return render_to_pdf(
        template_name,
        {
            "siswa": records,
            "jam": date_obj.strftime("%H:%M"),
            "tgl":hari(day)+', '+tanggal+' '+bulan(bln)+' '+tahun,
            "tgl1":date_obj.today().strftime("%d")+ ' '+bulan(date_obj.today().strftime("%m"))+' '+date_obj.today().strftime("%Y"),
            "romawi": romawi(bln),
            "tahun": tahun
        },
    )

def pengeluaran(request):
    if request.session.get("logged_in") != True:
        return redirect("login")
    template_name = "formpengeluaran.html"
    siswa = request.GET['siswa']
    records = Siswa.objects.filter(id_siswa=siswa).all()
    return render_to_pdf(
        template_name,
        {
            "siswa": records,
            "tgl":hari(datetime.today().strftime("%A")).lower()+' tanggal '+datetime.today().strftime("%d")+ ' '+bulan(datetime.today().strftime("%m"))+' '+datetime.today().strftime("%Y"),
            "tgl1":datetime.today().strftime("%d")+ ' '+bulan(datetime.today().strftime("%m"))+' '+datetime.today().strftime("%Y"),
            "tahun": datetime.today().strftime("%Y")
        },
    )



def render_to_pdf(template_src, context_dict={}):
    template = get_template(template_src)
    html = template.render(context_dict)
    response = HttpResponse(content_type='application/pdf')
    pdf_status = pisa.CreatePDF(html, dest=response)

    if pdf_status.err:
        return HttpResponse('Some errors were encountered <pre>' + html + '</pre>')

    return response


def hari(x):
    match x:
        case 'Sunday':
            hari ='Minggu'
        case 'Monday':
            hari ='Senin'
        case 'Tuesday':
            hari ='Selasa'
        case 'Wednesday':
            hari ='Rabu'
        case 'Thursday':
            hari ='Kamis'
        case 'Friday':
            hari ="Jum'at"
        case 'Saturday':
            hari ='Sabtu'
    return hari

def bulan(x):
    x = int(x)
    match x:
        case 1:
            bulan = 'Januari'
        case 2:
            bulan = 'Februari'
        case 3:
            bulan = 'Maret'
        case 4:
            bulan = 'April'
        case 5:
            bulan = 'Mei'
        case 6:
            bulan = 'Juni'
        case 7:
            bulan = 'Juli'
        case 8:
            bulan = 'Agustus'
        case 9:
            bulan = 'September'
        case 10:
            bulan = 'Oktober'
        case 11:
            bulan = 'November'
        case 12:
            bulan = 'Desember'
    return bulan

def romawi(x):
    x = int(x)
    match x:
        case 1:
            bulan = 'I'
        case 2:
            bulan = 'II'
        case 3:
            bulan = 'III'
        case 4:
            bulan = 'IV'
        case 5:
            bulan = 'V'
        case 6:
            bulan = 'VI'
        case 7:
            bulan = 'VII'
        case 8:
            bulan = 'VIII'
        case 9:
            bulan = 'IX'
        case 10:
            bulan = 'X'
        case 11:
            bulan = 'XI'
        case 12:
            bulan = 'XII'
    return bulan

@csrf_exempt
def impsiswa(request):
    if request.method == 'POST':
        excel_file = request.FILES['file']

        if excel_file.name.endswith('.xlsx'):
            wb = openpyxl.load_workbook(excel_file)
            sheet = wb.active

            for row in sheet.iter_rows(min_row=2, values_only=True):
                print(row)
                no,nama_lengkap, tempat, tgl_lahir, gender, alamat, lembaga, nomor_induk,jurusan,kelas,rombel = row
                try:
                    qr_code_url = nomor_induk
                    qr_code = qrcode.QRCode(
                        version=1,
                        error_correction=qrcode.constants.ERROR_CORRECT_L,
                        box_size=10,
                        border=4,
                    )
                    qr_code.add_data(qr_code_url)
                    qr_code.make(fit=True)
                    qr_code_image = qr_code.make_image(
                        fill_color="black", back_color="white"
                    )

                    qr_code_buffer = BytesIO()
                    qr_code_image.save(qr_code_buffer, format="PNG")

                    qr_code_directory = os.path.join("media", "siswa")
                    os.makedirs(qr_code_directory, exist_ok=True)

                    qr_code_filepath = os.path.join(
                        qr_code_directory, f"qr_code_{nomor_induk}.png"
                    )

                    with open(qr_code_filepath, "wb") as f:
                        f.write(qr_code_buffer.getvalue())
                    Siswa.objects.create(
                        nomor_induk=nomor_induk,
                        nama_lengkap=nama_lengkap,
                        tempat_lahir=tempat,
                        tanggal_lahir=tgl_lahir,
                        jenis_kelamin=gender,
                        alamat=alamat,
                        lembaga=lembaga,
                        jurusan=jurusan,
                        kelas=kelas,
                        rombel=rombel,
                        qr_code=f"qr_code_{nomor_induk}.png"
                    )
                    arr=1
                except IndexError:
                    arr:0
            arr = 1
    else:
        arr = 0
    return HttpResponse(arr)

@csrf_exempt
def export_laporan(request):
    tgl = request.POST.get("tgl")
    kategori = request.POST.get("kategori")
    siswa = request.POST.get("siswa")
    output = io.BytesIO()

    wb = xlsxwriter.Workbook(output)
    ws = wb.add_worksheet("Data Rekapan Pelanggaran")  # this will make a sheet named Users Data
    # Sheet header, first row
    date_format = '%Y-%m-%d %H:%M:%S'
    date_obj = datetime.strptime(tgl+"-01 00:00:00", date_format)
    # Merge 3 cells.
   
    row_num = 2
    cell_format = wb.add_format(
        {
            "bold": True,
            "font_color": "black",
            "align": "center",
            "valign": "vcenter",
            "border": 1,
        }
    )
    ws.merge_range("A1:J1", "REKAPAN LAPORAN PELANGGARAN SMK NURUL JADID", cell_format)
    ws.merge_range("A2:J2", "BULAN "+date_obj.strftime("%B %Y").upper(), cell_format)
    columns = [
        "No",
        "Nomor Induk",
        "Nama Siswa",
        "Kelas",
        "Pelanggaran",
        "Point",
        "Status Pelanggaran",
        "Total Point Pelanggaran",
        "Tindakan",
        "Tanggal Melanggar",
    ]

    for col_num in range(len(columns)):
        ws.write(row_num, col_num, columns[col_num], cell_format)  # at 0 row 0 column
        ws.set_column(col_num, col_num, len(columns[col_num]) + 10)

    style = wb.add_format(
        {
            "font_color": "black",
            "align": "center",
            "valign": "vcenter",
            "border": 1,
        }
    )
    with connection.cursor() as con:
        if kategori == 0:
            con.execute(
                "SELECT * FROM `siswa_siswapelanggaran` WHERE siswa_id=%s substr(tanggal_pelanggaran,1,7)=%s",
                [siswa, tgl],
            )
        else:
            con.execute(
                "SELECT * FROM `siswa_siswapelanggaran` WHERE substr(tanggal_pelanggaran,1,7)=%s ORDER BY siswa_id ASC,tanggal_pelanggaran ASC",
                [tgl],
            )
        rows = con.fetchall()
        no = 1
        for val in rows:
            row_num += 1
            sis = Siswa.objects.filter(id_siswa=val[2]).all()
            for v in sis.iterator():
                siswa = v.nama_lengkap.upper()
                kelas = v.kelas
                induk = v.nomor_induk
            pel = Pelanggaran.objects.filter(id_pelanggaran=val[1]).all()
            for v in pel.iterator():
                pelanggaran = v.nama_pelanggaran.upper()
                point = str(v.point_pelanggaran) + " Point"
                status = v.kategori.capitalize()
            ttl = SiswaPelanggaran.objects.filter(siswa_id=val[2]).all()
            total = 0
            for v in ttl.iterator():
                pel = Pelanggaran.objects.filter(id_pelanggaran=v.pelanggaran_id).all()
                for t in pel.iterator():
                    total += t.point_pelanggaran
            with connection.cursor() as kon:
                kon.execute("SELECT tindakan FROM siswa_tindakan WHERE range_point <=%s ORDER BY range_point DESC LIMIT 1",[total])
                row = kon.fetchall()
                if row != None:
                    for v in row:
                        tindakan = v[0].upper()
                else:
                    tindakan = "-"
            tanggal = val[3].strftime("%d/%m/%Y, %H:%M")
            value = []
            value = [no, induk, siswa,kelas, pelanggaran, point, status, str(total) + " Point",tindakan,tanggal]
            for col_num in range(len(value)):
                ws.write(row_num, col_num, value[col_num], style)
                ws.set_column(col_num, col_num, len(columns[col_num]) + 10)
            no += 1
    wb.close()
    # Rewind the buffer.
    output.seek(0)
    # Set up the Http response.
    filename = "Pelanggaran.xlsx"
    response = HttpResponse(
        output,
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = "attachment; filename=%s" % filename
    return response

@csrf_exempt
def chartpelanggaran(request):
    year = str(request.POST.get("tahun"))
    count = 1
    data = []
    while count <= 12:
        if count < 10:
            bln = year + "-0" + str(count)
        else:
            bln = year + "-" + str(count)
        with connection.cursor() as con:
            con.execute(
                "SELECT COUNT(*) FROM siswa_siswapelanggaran WHERE substr(tanggal_pelanggaran,1,7)=%s",
                [bln],
            )
            rows = con.fetchone()
            jumlah = int(rows[0])
            if rows[0] % 1000 == 0:
                hsl = rows[0] / 1000
                if len(str(hsl)) > 3:
                    hrg = f"{hsl:,}" + "k"
                else:
                    hrg = str(hsl) + "k"
            else:
                hrg = str(rows[0])+"x"
            harga = "(" + hrg + ")"
        out = [jumlah,harga]
        data.append(out)
        count += 1
    output = json.dumps({"data": data})
    return HttpResponse(output, content_type="application/json")